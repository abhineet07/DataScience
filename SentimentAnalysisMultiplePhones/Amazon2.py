from bs4 import BeautifulSoup as soup, NavigableString, Tag
import requests
import re
from Trie import *
from ToJSON import *
from ExcelWriter import *
from GoogleToAmazon import *
# from SentimentAnalysis2 import *
from SantimentAnalysis3 import *
from openpyxl import Workbook, load_workbook
import time
import os
import urllib.request
from StyledExcel import *
import csv

class AmazonClass:
    def __init__(self, brand, model):
        self.brand = brand
        self.model = model

    def seachByName(self, query, pages):
        google = Google()
        # url = google.search(query)[0]
        ######################################
        urls = google.searchn(query, 6)
        url = ""
        for link in urls:
            # print(link)
            str_link = str(link)
            if str_link.__contains__("dp"):
                url = link
                break

        temp_url = str(url).split("/")
        url = "http://www.amazon.in/product-reviews/" + temp_url[-1]
        # url = "https://www.amazon.in/Vivo-Mineral-Storage-Additional-Exchange/product-reviews/B07KXC5H3N/ref=cm_cr_dp_d_show_all_top?ie=UTF8&reviewerType=all_reviews"
        print(url)
        self.postURL(url, pages)

    def searchByUrl(self, url, pages):
        # https://www.amazon.in/Vivo-Frozen-Storage-Additional-Exchange/dp/B07KXCGR8Q/ref=cm_cr_arp_d_product_top?ie=UTF8
        req_page = requests.get(url)
        print(url)
        print("STATUS CODE : ", req_page.status_code)
        page_html = req_page.text
        page_soup = soup(page_html, 'html.parser')
        # print(page_soup)
        review_page_link = page_soup.find("a", {"id": "dp-summary-see-all-reviews"})
        print(review_page_link)
        new_page_url = "http://www.amazon.in" + str(review_page_link["href"])
        print(new_page_url)
        self.postURL(new_page_url, pages)

    # def postURL(self, url, pages):
    #     print(url)
    #     self.data = []
    #     page_soup = self.getPageSoup(url)
    #     # print(page_soup)
    #     self.analysis(page_soup)
    #
    #     prev_url = url
    #     for p in range(1, pages):
    #         time.sleep(2)
    #         li_soup = page_soup.find("li", {"class": "a-last"})
    #         a_soup = soup(str(li_soup), 'html.parser').find("a")
    #         # print(a_soup["href"])
    #
    #         while a_soup == None:
    #             print("Reloading same page")
    #             time.sleep(2)
    #             page_soup = self.getPageSoup(prev_url)
    #             li_soup = page_soup.find("li", {"class": "a-last"})
    #             a_soup = soup(str(li_soup), 'html.parser').find("a")
    #             # print(a_soup["href"])
    #
    #         prev_url = url
    #         url = str("http://www.amazon.in" + str(a_soup["href"]))
    #         print(url)
    #         page_soup = self.getPageSoup(url)
    #         self.analysis(page_soup)
    #
    #     self.writeToExcel()  # writing data to excel

    # def postURL(self, url, pages):
    #     self.data = []
    #     for i in range(1, pages):
    #         next_page_soup = None
    #         a_soup = None
    #         reload_value = 0
    #         while True:
    #             print("Reloading Value : ", reload_value, ' URL : ', url)
    #             page_soup = self.getPageSoup(url)
    #             try:
    #                 next_page_soup = soup.find("li", {"class": "a-disabled a-last"})
    #             except:
    #                 next_page_soup = None
    #
    #             if next_page_soup != None:
    #                 break
    #             li_soup = page_soup.find("li", {"class": "a-last"})
    #
    #             a_soup = soup(str(li_soup), 'html.parser').find("a")
    #             reload_value += 1
    #
    #         self.analysis(page_soup)
    #
    #         url = str("http://www.amazon.in" + str(a_soup["href"]))
    #
    #     self.writeToExcel()  # writing data to excel

    def postURL(self, url, pages):
        self.data = []
        for i in range(1, pages):
            next_page_soup = None
            page_soup = self.getPageSoup(url)
            try:
                next_page_soup = page_soup.find("li", {"class": "a-disabled a-last"})
            except:
                next_page_soup = None

            if next_page_soup != None:
                print("LAST PAGE HURRAYYYY....")
                break

            li_soup = page_soup.find("li", {"class": "a-last"})
            a_soup = soup(str(li_soup), 'html.parser').find("a")

            self.analysis(page_soup)

            url = str("http://www.amazon.in" + str(a_soup["href"]))

        self.writeToExcel()  # writing data to excel

    def getPageSoup(self, url):
        page = requests.get(url)
        time.sleep(15)
        status_code = page.status_code
        print("INSIDE getPageSoup()")
        print("\tURL : ", url)
        print("\tSTATUS CODE : ", status_code)
        reload = 1
        while status_code > 200:
            time.sleep(15)
            page = requests.get(url)
            status_code = page.status_code
            print("\tStatus Code : ", status_code, " :: RELOAD : ", reload)
            reload += 1
        page_html = page.text
        page_soup = soup(page_html, "html.parser")

        focus_area = page_soup.find("div", {"id": "cm_cr-review_list"})
        page_soup = soup(str(focus_area), 'html.parser')
        # print("page soup retrieved")
        print("Returned from page soup")
        return page_soup


    def analysis(self, page_soup):
        # Review Title
        title_list = []
        title_tags = page_soup.findAll("a", {"data-hook": "review-title"})
        for tt in title_tags:
            curr_title = str(tt.text).strip()
            curr_title = curr_title.encode('ascii', 'ignore').decode('ascii')
            title_list.append(curr_title)

            # title_list.append(str(tt.text).strip())
            # print(str(tt.text).strip())

        # Review Date
        date_list = []
        date_tags = page_soup.findAll("span", {"data-hook": "review-date"})
        for dt in date_tags:
            date_list.append(str(dt.text).strip())

        # Ratings
        rating_list = []
        rating_tags = page_soup.findAll("span", {"class": "a-icon-alt"})
        for rt in rating_tags:
            rate = str(rt.text).split(" ")[0] + "/5"
            rating_list.append(rate)

        # Comment
        comment_list = []
        comment_tags = page_soup.findAll("span", {"data-hook": "review-body"})
        for ct in comment_tags:
            cmt_soup = soup(str(ct), 'html.parser')
            for br in cmt_soup.findAll("br"):
                br.replace_with(" ")

            curr_cmt = str(cmt_soup.text).strip()
            curr_cmt = curr_cmt.encode('ascii', 'ignore').decode('ascii')
            comment_list.append(curr_cmt)

            # comment_list.append(str(cmt_soup.text).strip())

        sn = Sentiment()
        i = 0
        for cmt in comment_list:
            sentiment_value, all_features_set, positive_feature_set, positive_points, negative_feature_set, negative_points = sn.sentimentAnalysis(str(cmt))
            all_features_set_str = str(",".join(x for x in all_features_set))
            positive_feature_set_str = str(",".join(x for x in positive_feature_set))
            positive_points_str = str(",".join(x for x in positive_points))
            negative_feature_set_str = str(",".join(x for x in negative_feature_set))
            negative_points_str = str(",".join(x for x in negative_points))
            self.data.append(["Amazon", self.brand, self.model, title_list[i], date_list[i], rating_list[i],sentiment_value, all_features_set_str,
                              positive_feature_set_str, positive_points_str,
                              negative_feature_set_str, negative_points_str, cmt])
            i += 1

        # for d in self.data:
        #     print(d)

    def writeToExcel(self):
        filename1 = "Amazon_" + self.brand + "_" + self.model + ".xlsx"
        filename2 = "Reviews_" + self.brand + "_" + self.model + ".xlsx"
        filename3 = open("Amazon_" + self.brand + "_" + self.model + ".csv", 'a')
        writer = csv.writer(filename3)


        if not os.path.isfile(filename1):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "Title", "Date", "User Rating", "Sentiment Value", "All Features Apperared",
                       "Positive Feature Set", "Positive Points",
                       "Negative Feature Set", "Negative Points", "Comment"])

            wb.save(filename1)

        if not os.path.isfile(filename2):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(
                ["Source", "Brand", "Model", "Title", "Date", "User Rating", "Sentiment Value", "All Features Apperared",
                 "Positive Feature Set", "Positive Points",
                 "Negative Feature Set", "Negative Points", "Comment"])
            wb.save(filename2)


        # data = ["Flipkart", self.brand, self.model, , self.pros_list, self.cons_list]

        wb1 = load_workbook(filename1)
        wb2 = load_workbook(filename2)
        ws1 = wb1.worksheets[0]
        ws2 = wb2.worksheets[0]

        # for d in self.data:
        #     print(d)

        for d in self.data:
            writer.writerow(d)

        for d in self.data:
            # print(d)
            ws1.append(d)
            ws2.append(d)
        wb1.save(filename1)
        wb2.save(filename2)

        newfilename = 'UserReviewsData_' + self.model + '.xlsx'
        se = styledExcel(newfilename)
        se.makeData(filename2)
        print(">>> Data added and saved... <<<")


# amz = AmazonClass("samsung", "a10")
# amz.searchByUrl("https://www.amazon.in/Samsung-Galaxy-Charcoal-Black-64GB/dp/B07HGJFSC9?tag=googinkenshoo-21&ascsubtag=_k_CjwKCAjwsIbpBRBNEiwAZF8-zx8ZXA7t5Tgo44f4S8Khzsd8qDW7O0NqMVq7eD7tUv0J_YAqhkKN2BoC7qYQAvD_BwE_k_&gclid=CjwKCAjwsIbpBRBNEiwAZF8-zx8ZXA7t5Tgo44f4S8Khzsd8qDW7O0NqMVq7eD7tUv0J_YAqhkKN2BoC7qYQAvD_BwE", 50)

# vivo y12
# amz = AmazonClass("vivo", "y12")
# url = "https://www.amazon.in/Vivo-Storage-Additional-Exchange-Offers/dp/B07PTTMDM4/ref=cm_cr_arp_d_product_top?ie=UTF8"
# amz.searchByUrl(url, 1000)