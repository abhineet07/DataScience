import requests
import urllib3
# from requests.packages.urllib3.exceptions import InsecureRequestWarning
from GoogleToAmazon import *
from GoogleToAmazon2 import *
from bs4 import BeautifulSoup as soup, NavigableString, Tag
from textblob import TextBlob
import re
from Trie import *
import os
from openpyxl import Workbook, load_workbook
from StyledExcel import *

# vivo v15 pro 91mobiles hub review

class Mobiles91:
    def __init__(self):
        self.pros_list = ""
        self.cons_list = ""

    def searchByName(self, query, brand, model):
        google = Google()
        url = google.search(query)[0]
        # print(url)
        self.searchByURL(url, brand, model)

    def searchByURL(self, url, brand, model):
        print(url)
        page_soup = self.getPageSoup(url)
        self.analysis(page_soup)
        self.writeToExcel(url, brand, model)

    def getPageSoup(self, url):
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        page = requests.get(url, verify=False)
        page_html = page.text
        # print(page_html)
        page_soup = soup(page_html, "html.parser")

        return page_soup

    def analysis(self, page_soup):
        # div_soup = page_soup.findAll("li", {"style": "font-size: 14px; color: #333; font-weight: 300; margin: 5px 0; line-height: 15px;"})
        positive_list = []
        negative_list = []

        div_soup = page_soup.findAll("ul", {
            "style": "margin: 0px 6px 7px 20px; font-size: 15px; line-height: 20px; color: #414141; list-style: disc; padding-left: 0px;"})

        pros_soup = soup(str(div_soup[0]), 'html.parser').findAll("li")
        for p in pros_soup:
            positive_list.append("- " + p.text)

        cons_soup = soup(str(div_soup[1]), 'html.parser').findAll("li")
        for c in cons_soup:
            negative_list.append("- " + p.text)

        self.pros_list = str("\n".join(s for s in positive_list))
        self.cons_list = str("\n".join(s for s in negative_list))


    def writeToExcel(self, url, brand, model):
        if not os.path.isfile("AnalysisReport.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "URL", "Positives", "Negatives"])
            wb.save("AnalysisReport.xlsx")

        data = ["Mobile91", brand, model, str(url), self.pros_list, self.cons_list]
        wb = load_workbook("AnalysisReport.xlsx")
        ws = wb.worksheets[0]
        ws.append(data)
        wb.save("AnalysisReport.xlsx")
        se = styledExcel('CriticsData.xlsx')
        se.makeData('AnalysisReport.xlsx')
        print(">>> Data added and saved... <<<")


# mo = Mobiles91("vivo v15 pro 91mobiles hub review")
# mo = Mobiles91()
# mo.searchByName("vivo v15 pro 91mobiles hub review", "vivo", "v15 pro")