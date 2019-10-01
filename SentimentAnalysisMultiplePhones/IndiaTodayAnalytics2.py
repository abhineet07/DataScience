import requests
import urllib3
# from requests.packages.urllib3.exceptions import InsecureRequestWarning
from GoogleToAmazon import *
from GoogleToAmazon2 import *
from bs4 import BeautifulSoup as soup, NavigableString, Tag
from textblob import TextBlob
import re
from Trie import *
from openpyxl import Workbook, load_workbook
import os
from StyledExcel import *

# samsung a50 review india today group

class IndiaToday:
    def __init__(self):
        self.pros_list = ""
        self.cons_list = ""

    def seacrchByName(self, query, brand, model):
        google = Google()
        url = google.search(query)[0]
        # print(url)
        self.searchByURL(url, brand, model)

    def searchByURL(self, url, brand, model):
        page_soup = self.getPageSoup(url)
        self.analysis(page_soup)

        self.writeToExcel(url, brand, model)

    def getPageSoup(self, url):
        # requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        page = requests.get(url, verify=False)
        page_html = page.text
        # print(page_html)
        page_soup = soup(page_html, "html.parser")

        return page_soup

    def analysis(self, page_soup):
        positive_list = []
        neagtive_list = []

        # reviewPro || tech-pros
        div_pro = page_soup.find("div", {"class": "reviewPro"})
        pros = soup(str(div_pro), 'html.parser').findAll("li")
        for p in pros:
            positive_list.append("- " + p.text)

        div_cons = page_soup.find("div", {"class": "reviewCons"})
        cons = soup(str(div_cons), 'html.parser').findAll("li")
        for c in cons:
            neagtive_list.append("- " + c.text)

        self.pros_list = str("\n".join(s for s in positive_list))
        self.cons_list = str("\n".join(s for s in neagtive_list))


    def writeToExcel(self, url, brand, model):
        if not os.path.isfile("AnalysisReport.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "URL", "Positives", "Negatives"])
            wb.save("AnalysisReport.xlsx")

        data = ["IndiaToday", brand, model, str(url), self.pros_list, self.cons_list]
        wb = load_workbook("AnalysisReport.xlsx")
        ws = wb.worksheets[0]
        ws.append(data)
        wb.save("AnalysisReport.xlsx")
        se = styledExcel('CriticsData.xlsx')
        se.makeData('AnalysisReport.xlsx')
        print(">>> Data added and saved... <<<")


# ind = IndiaToday()
# ind.seacrchByName("samsung a50 review india today group", "samsung", "a50")