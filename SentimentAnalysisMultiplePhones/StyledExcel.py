from openpyxl import *
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
import re
import os

class styledExcel:
    def __init__(self, newfilename):
        self.newfilename = newfilename
        if newfilename=='CriticsData.xlsx' and os.path.exists('CriticsData.xlsx'):
            os.remove('CriticsData.xlsx')
        elif newfilename=='UserReviewsData.xlsx' and os.path.exists('UserReviewsData.xlsx'):
            os.remove('UserReviewsData.xlsx')

    def makeData(self, filename1):
        wb1 = load_workbook(filename1)
        ws1 = wb1.worksheets[0]

        # new excel file
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Analysis"

        # style for headings
        heading_style = NamedStyle(name="heading_style")
        heading_style.font = Font(bold=True, size=15, color='FF000000')
        heading_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        heading_style.fill = PatternFill(patternType='solid', start_color='B2E6F6', end_color='B2E6F6')
        wb2.add_named_style(heading_style)

        # style for rest of the data
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(bold=True, color='FF000000')
        data_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        wb2.add_named_style(data_style)

        # columns with less width
        short_cols = []
        if self.newfilename == 'CriticsData.xlsx':
            short_cols = ['A', 'B', 'C']
        elif self.newfilename == 'UserReviewsData.xlsx':
            short_cols = ['A', 'B', 'C', 'E', 'F']

        # setting columns width
        for col in ws1.columns:
            col_name = re.findall('\w\d', str(col[0]))
            col_name = col_name[0]
            col_name = re.findall('\w', str(col_name))[0]
            if col_name in short_cols:
                ws2.column_dimensions[col_name].width = 20
            else:
                ws2.column_dimensions[col_name].width = 40

        # setting columns height
        for row in ws1.rows:
            row_name = int(re.findall('\d+', str(row[0]))[-1])
            ws2.row_dimensions[row_name].height = 70

        # traversing each cell
        index = 0
        for row in ws1.rows:
            for cell in row:
                cell_name = re.findall('\w+\d+', str(cell))[0]
                ws2[cell_name].value = ws1[cell_name].value
                if index == 0:
                    ws2[cell_name].style = heading_style
                else:
                    ws2[cell_name].style = data_style
            index += 1


        wb2.save(self.newfilename)

    def makeUserReviewData(self, filename1):
        wb1 = load_workbook(filename1)
        ws1 = wb1.worksheets[0]

        # new excel file
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Analysis"


# se = styledExcel('CriticsData.xlsx')
# se.makeData('AnalysisReport.xlsx')

# se = styledExcel('UserReviewsData.xlsx')
# se.makeData('Reviews_Samsung_a50.xlsx')