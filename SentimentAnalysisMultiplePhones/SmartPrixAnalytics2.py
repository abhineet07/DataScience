import requests
from GoogleToAmazon import *
from GoogleToAmazon2 import *
from bs4 import BeautifulSoup as soup, NavigableString, Tag
from textblob import TextBlob
import re
from Trie import *
from ExcelWriter import *
import os
from openpyxl import Workbook, load_workbook
from StyledExcel import *

class SmartPrix:
    def __init__(self):
        self.pros_list = ""
        self.cons_list = ""

    def searchByName(self, query, brand, model):
        google = Google()
        url = google.search(query)[0]
        # print(url)
        self.searchByURL(url, brand, model)


    def searchByURL(self, url, brand, model):
        page_soup = self.getPageSoup(url)
        self.analysis(page_soup)
        self.writeToExcel(url, brand, model)

    def getPageSoup(self, url):
        page = requests.get(url)
        page_html = page.text
        page_soup = soup(page_html, "html.parser")
        return page_soup

    def analysis(self, page_soup):
        positive_list = []
        negative_list = []

        div = page_soup.find("div", {'class' : 'procons'})
        div_soup = soup(str(div), 'html.parser')

        pros_i = div_soup.findAll("ul", {"class" : "pros"})
        pros_i = soup(str(pros_i), 'html.parser').findAll("li")
        for i in pros_i:
            # print(i.text)
            positive_list.append("- " + i.text)
            # self.pros_list.append(i.text)


        cons_i = div_soup.findAll("ul", {"class": "cons"})
        cons_i = soup(str(cons_i), 'html.parser').findAll("li")
        for i in cons_i:
            # print(i.text)
            negative_list.append("- " + i.text)
            # self.cons_list.append(i)

        self.pros_list = str("\n".join(s for s in positive_list))
        self.cons_list = str("\n".join(s for s in negative_list))


    def writeToExcel(self, url, brand, model):
        if not os.path.isfile("AnalysisReport.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "URL", "Positives", "Negatives"])
            wb.save("AnalysisReport.xlsx")

        data = ["SmartPrix", brand, model, str(url), self.pros_list, self.cons_list]
        wb = load_workbook("AnalysisReport.xlsx")
        ws = wb.worksheets[0]
        ws.append(data)
        wb.save("AnalysisReport.xlsx")
        se = styledExcel('CriticsData.xlsx')
        se.makeData('AnalysisReport.xlsx')
        print(">>> Data added and saved... <<<")


#### HANDLER ####
# sm = SmartPrix("smartprix samsung a50 specs and reviews")

# sm = SmartPrix()
# sm.searchByName("smartprix samsung a50 specs and reviews", "Samsung", "a50")
