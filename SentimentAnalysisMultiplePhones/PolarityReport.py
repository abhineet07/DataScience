from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
import pandas as pd
import re
import os

class polarityReport:
    def __init__(self, filename, sheetname, modelname):
        self.fname = filename
        self.sname = sheetname
        self.modelname = modelname
        self.dataframe = dfs = pd.read_excel(filename, sheet_name=sheetname)  # <type:dataframe>
        rows, cols = self.dataframe.shape
        self.dictionary = {"phone": ["phone", "device"],
                           "camera": ["camera", "photos", "picture"],
                           "battery": ["battery", "backup"],
                           "display": ["display", "screen", "notch", "amoled"],
                           "finger": ["finger"],
                           "price": ["price", "cost", "value for money", "budget", "affordable"],
                           "performance": ["performance", "perform", "sluggish", "lag", "hang"],
                           "face": ["face", "face unlock", "facial"],
                           "charging": ["charge", "charging", "charger", "vooc"],
                           "design": ["design", "look", "build"],
                           "call": ["call"],
                           "heating": ["heat", "heating"],
                           "game": ["game", "pubg"],
                           "speaker": ["speaker", "sound", "headphones"],
                           "ram": ["ram", "memory"],
                           "bloatware": ["bloatware"],
                           "ui": ["ui", "one ui", "one-ui"]}

    def analyseReport(self):
        data = []
        wb = Workbook()
        ws = wb.active
        ws.title = "Polarity"
        ws.append(["Category", "Positive Polarity", "Positive", "Negative"])
        for k, v in self.dictionary.items():
            # print("Key : ", k, "\t Values : " , v)
            data.append(self.calculatePolarity(k))
            ws.append(self.calculatePolarity(k))



        for d in data:
            print(d)

        # heading style
        heading_style = NamedStyle(name="heading_style")
        heading_style.font = Font(bold=True, color='FF000000')
        heading_style.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        heading_style.fill = PatternFill(patternType='solid', start_color='FFF300', end_color='FFF300')
        heading_style.border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        wb.add_named_style(heading_style)

        # first column style
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(bold=True, color='FF000000')
        data_style.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        data_style.fill = PatternFill(patternType='solid', start_color='FFBD4A', end_color='FFBD4A')
        data_style.border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        wb.add_named_style(data_style)

        # cell style
        common_style = NamedStyle(name="common_style")
        common_style.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        common_style.border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        wb.add_named_style(common_style)

        # setting column width
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # applying style
        i, j = 0, 0
        for row in ws.rows:
            j = 0
            for cell in row:
                cell_name = re.findall('\w+\d+', str(cell))[0]
                if i == 0:
                    ws[cell_name].style = heading_style
                elif i > 0 and j == 0:
                    ws[cell_name].style = data_style
                else:
                    ws[cell_name].style = common_style
                j += 1
            i += 1

        newfilename = "Polarity_Analysis_" + self.modelname + ".xlsx"
        wb.save(newfilename)
        os.startfile(newfilename)

    def calculatePolarity(self, feature):
        relative_features = self.dictionary[feature]
        all_count = 0

        all_features = self.dataframe["All Features Apperared"]
        for af in all_features.dropna():
            for rf in relative_features:
                if af.__contains__(rf):
                    all_count += 1
                    break
        # print("Total Occurance : ", all_count)

        positive_count = 0
        all_features = self.dataframe["Positive Feature Set"]
        for af in all_features.dropna():
            for rf in relative_features:
                if af.__contains__(feature):
                    positive_count += 1
                    break
        # print("Positive : ", positive_count, "\t", (positive_count / all_count) * 100, "%")
        # print("Positive : ", positive_count)

        negative_count = 0
        all_features = self.dataframe["Negative Feature Set"]
        for af in all_features.dropna():
            if af.__contains__(feature):
                for rf in relative_features:
                    negative_count += 1
                    break
        # print("Neagtive : ", negative_count, "\t", (negative_count / all_count) * 100, "%")
        # print("Neagtive : ", negative_count)

        pos_polarity = 0
        try:
            pos_polarity = round((positive_count / (positive_count+negative_count)) * 100, 2)
            pos_polarity = str(pos_polarity) + "%"
        except:
            pos_polarity = '0%'
        return [feature, pos_polarity, positive_count, negative_count]

    def getPolarityDictionary(self, polarity):
        my_dict = {}

        j = 2
        if polarity == 'Positive':
            j = 2
        elif polarity == 'Negative':
            j = 3

        for k, v in self.dictionary.items():
            temp_list = self.calculatePolarity(k)
            my_dict[temp_list[0]] = int(temp_list[j])

        # print(my_dict)
        return my_dict


# pr = polarityReport('Reviews_Samsung_m10.xlsx', 'Analysis', 'm10')
# pr.getPolarityDictionary("Positive")