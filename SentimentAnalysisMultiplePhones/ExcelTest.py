from openpyxl import *
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
import re

filename1 = 'AnalysisReport.xlsx'
wb = load_workbook(filename1)
ws = wb.worksheets[0]

# SETTING COLUMN DIMENSION
for col in ws.columns:
    # print("Step-0 : ", col[0])
    col_name = re.findall('\w\d', str(col[0]))
    # print("Step-1 : ", col_name)
    col_name = col_name[0]
    # print("Step-2 : ", col_name)
    col_name = re.findall('\w', str(col_name))[0]
    # print("Step-3 : ", col_name)
    ws.column_dimensions[col_name].width = 40

# SETTING HEIGHT
for row in ws.rows:
    # print(row[0])
    row_name = int(re.findall('\d+', str(row[0]))[-1])
    # print(row_name, type(row_name))
    ws.row_dimensions[row_name].height = 70

# Styling
# highlight = NamedStyle(name="highlight")
# highlight.font = Font(bold=True, color='FF000000')
# highlight.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
# highlight.fill = PatternFill(patternType='solid', start_color='B2E6F6', end_color='B2E6F6')
# wb.add_named_style(highlight)

# GETTING EACH CELL NAME
for row in ws.rows:
    for cell in row:
        cell_name = re.findall('\w+\d+', str(cell))[0]
        print(ws[cell_name].value)
        # ws[cell_name].style = highlight



# wb.save(filename1)

