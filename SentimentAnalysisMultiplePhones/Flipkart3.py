from bs4 import BeautifulSoup as soup, NavigableString, Tag
import requests
import re
from Trie import *
from ToJSON import *
from ExcelWriter import *
from GoogleToAmazon import *
# from SentimentAnalysis import *
from SantimentAnalysis3 import *
from openpyxl import Workbook, load_workbook
import time
import os
from StyledExcel import *

class Flipkart:
    def __init__(self, brand, model):
        self.brand = brand
        self.model = model

    def seachByUrl(self, url, pages):
        # https://www.flipkart.com/samsung-galaxy-a50-black-64-gb/p/itmfe4csknzx2kcs
        req_page = requests.get(url)
        page_html = req_page.text
        page_soup = soup(page_html, 'html.parser')
        div_tag = page_soup.find("div", {"class": "swINJg _3nrCtb"})
        a_tag = div_tag.find_parent('a')

        url = "http://www.flipkart.com" + str(a_tag["href"])
        print(url)
        self.postURL(url, pages)

    def searchByName(self, query, pages):
        google = Google()
        url = google.search(query)[0]
        print("url : ", url)
        self.postURL(url, pages)

    # def postURL(self, url, pages):
    #     print(url)
    #     page_soup = self.getPageSoup(url)
    #     # self.ratingAnalysis(page_soup)
    #     self.data = []
    #     self.analysis(page_soup)
    #
    #     for p in range(1, pages):
    #         time.sleep(2)
    #         next_page = page_soup.findAll("a", {"class": "_3fVaIS"})[-1]
    #         url = "https://www.flipkart.com" + str(next_page["href"])
    #         print(url)
    #         page_soup = self.getPageSoup(url)
    #         self.analysis(page_soup)
    #
    #     self.writeToExcel() # writing data to excel

    def postURL(self, url, pages):
        print("Post url : ", url)
        self.data = []
        for i in range(1, pages):
            next_page = None
            reload_value = 0
            while next_page == None:
                print("Reloading Value : ", reload_value, ' URL : ', url)

                page_soup = self.getPageSoup(url)
                next_page = page_soup.findAll("a", {"class": "_3fVaIS"})[-1]
                reload_value += 1

            self.analysis(page_soup)
            url = "https://www.flipkart.com" + str(next_page["href"])

        self.writeToExcel()  # writing data to excel


    def getPageSoup(self, url):
        status_code = 503
        page = requests.get(url)
        while status_code <= 200:
            page = requests.get(url)
            status_code = page.status_code
            time.sleep(15)
            print("Status Code : ", status_code)
        page_html = page.text
        page_soup = soup(page_html, "html.parser")
        return page_soup

    def isLastPage(self):
        pass

    def ratingAnalysis(self, page_soup):
        feature_rating_table = []
        new_links = []
        tab_links = page_soup.findAll("div", {"class": "_36Dmoj"})
        new_soup = soup(str(tab_links), 'html.parser').findAll("a")
        for ns in new_soup:
            new_links.append("https://www.flipkart.com" + ns["href"])

        for i in range(1, len(new_links)):
            temp_soup = self.getPageSoup(new_links[i])
            # FEATURE NAME
            curr_feature = temp_soup.find("span", {"class": "nR6alz _4C7vqx"})
            curr_feature_positive = re.findall(r'\d+', temp_soup.find("span", {"class": "_1Iw1jf"}).text)[0] + "%"
            curr_feature_negative = re.findall(r'\d+', temp_soup.find("span", {"class": "_3MuE_5"}).text)[0] + "%"

            feature_rating_table.append([str(curr_feature.text),
                                         curr_feature_positive,
                                         curr_feature_negative])

        print(*feature_rating_table, sep='\n')

    def analysis(self, page_soup):
        title_list = []
        date_list = []
        comments_list = []
        rating_list = []
        focus_area = page_soup.findAll("div", {"class" : "col _390CkK _1gY8H-"})
        for box in focus_area:
            new_soup = soup(str(box), 'html.parser')

            # TITLE
            titles = new_soup.findAll("p", {"class": "_2xg6Ul"})
            for t in titles:
                title_list.append(t.text)

            # DATE LIST
            dates = new_soup.findAll("p", {"class": "_3LYOAd"})
            ind = 1
            for d in dates:
                if ind % 2 == 0:
                    date_list.append(d.text)
                    # print(d.text)
                ind += 1

            # COMMENT
            comments = new_soup.findAll("div", {"class": "qwjRop"})
            for c in comments:
                cmt_soup = soup(str(c), 'html.parser')
                for br in cmt_soup.findAll("br"):
                    br.replace_with(" ")

                comment_text = str(cmt_soup.text)
                if comment_text.__contains__("READ MORE"):
                    comment_text = comment_text.replace("READ MORE", '')
                    # print(comment_text)
                comments_list.append(comment_text)

            # RATINGS
            ratings_row = new_soup.find("div", {"class": "row"})
            ratings = soup(str(ratings_row), 'html.parser').find_all("div")
            rating_list.append(ratings[1].text)

        # print("Title List length : ", len(title_list))
        # print("Comment List Length : ", len(comments_list))
        # print("Rating List Length : ", len(rating_list))

        sn = Sentiment()
        i = 0
        for cmt in comments_list:
            # print(cmt)
            sentiment_value, all_features_set, positive_feature_set, positive_points, negative_feature_set, negative_points = sn.sentimentAnalysis(str(cmt))
            all_features_set_str = str(",".join(x for x in all_features_set))
            positive_feature_set_str = str(",".join(x for x in positive_feature_set))
            positive_points_str = str(",".join(x for x in positive_points))
            negative_feature_set_str = str(",".join(x for x in negative_feature_set))
            negative_points_str = str(",".join(x for x in negative_points))
            self.data.append(["Flipkart", self.brand, self.model, title_list[i], date_list[i], rating_list[i],sentiment_value, all_features_set_str,
                              positive_feature_set_str, positive_points_str,
                              negative_feature_set_str, negative_points_str,cmt])
            i += 1

    def writeToExcel(self):
        filename1 = "Flipkart_" + self.brand + "_" + self.model + ".xlsx"
        filename2 = "Reviews_" + self.brand + "_" + self.model + ".xlsx"
        if not os.path.isfile(filename1):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "Title", "Date", "User Rating", "Sentiment Value", "All Features Apperared",
                       "Positive Feature Set", "Positive Points",
                       "Negative Feature Set", "Negative Points", "Comment"])
            wb.save(filename1)

        if not os.path.isfile(filename2):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(
                ["Source", "Brand", "Model", "Title", "Date", "User Rating", "Sentiment Value", "All Features Apperared",
                 "Positive Feature Set", "Positive Points",
                 "Negative Feature Set", "Negative Points", "Comment"])
            wb.save(filename2)

        # data = ["Flipkart", self.brand, self.model, , self.pros_list, self.cons_list]

        wb1 = load_workbook(filename1)
        wb2 = load_workbook(filename2)
        ws1 = wb1.worksheets[0]
        ws2 = wb2.worksheets[0]
        for d in self.data:
            ws1.append(d)
            ws2.append(d)
        wb1.save(filename1)
        wb2.save(filename2)

        newfilename = "UserReviewsData_" + self.model + ".xlsx"
        se = styledExcel(newfilename)
        se.makeData(filename2)
        print(">>> Data added and saved... <<<")
        #
        # for d in self.data:
        #     print(d)

# query = "samsung a50 flipkart product reviews"
# fp = Flipkart(query, 4, "samsung", "a50")

# fp = Flipkart("samsung", "a50")
# fp.seachByUrl("https://www.flipkart.com/samsung-galaxy-a50-black-64-gb/p/itmfe4csknzx2kcs", 3)


fp = Flipkart("vivo", "y17")
# this page contains "See all ### reviews"
fp.seachByUrl("https://www.flipkart.com/vivo-y17-mineral-blue-128-gb/p/itmffftgdyezevvg", 5)