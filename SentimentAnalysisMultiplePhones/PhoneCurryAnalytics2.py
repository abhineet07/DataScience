import requests
import urllib3
# from requests.packages.urllib3.exceptions import InsecureRequestWarning
from GoogleToAmazon import *
from GoogleToAmazon2 import *
from bs4 import BeautifulSoup as soup, NavigableString, Tag
from textblob import TextBlob
import re
from Trie import *
import os
from openpyxl import Workbook, load_workbook
from StyledExcel import *

class PhoneCurry:
    def __init__(self):
        self.pros_list = ""
        self.cons_list = ""

    def searchByName(self, query, brand, model):
        google = Google()
        url = google.search(query)[0]
        self.searchByURL(url, brand, model)

    def searchByURL(self, url, brand, model):
        print(url)
        page_soup = self.getPageSoup(url)
        self.analysis(page_soup)
        self.writeToExcel(url, brand, model)

    def getPageSoup(self, url):
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        page = requests.get(url, verify=False)
        page_html = page.text
        page_soup = soup(page_html, "html.parser")
        return page_soup

    def analysis(self, page_soup):
        positive_list = []
        negative_list = []
        pros = page_soup.findAll("li", {"class": "positive"})
        # print("PROS : ")
        for p in pros:
            bold = soup(str(p), 'html.parser').find("b")
            # print(bold.text)
            positive_list.append("- " + bold.text)
        print()
        # print("CONS : ")
        cons = page_soup.findAll("li", {"class": "negative"})
        for c in cons:
            bold = soup(str(c), 'html.parser').find("b")
            # print(bold.text)
            negative_list.append("- " + bold.text)

        self.pros_list = str("\n".join(s for s in positive_list))
        self.cons_list = str("\n".join(s for s in negative_list))

    def writeToExcel(self, url, brand, model):
        if not os.path.isfile("AnalysisReport.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Analysis"
            ws.append(["Source", "Brand", "Model", "URL", "Positives", "Negatives"])
            wb.save("AnalysisReport.xlsx")

        data = ["PhoneCurry", brand, model, str(url), self.pros_list, self.cons_list]
        wb = load_workbook("AnalysisReport.xlsx")
        ws = wb.worksheets[0]
        ws.append(data)
        wb.save("AnalysisReport.xlsx")
        se = styledExcel('CriticsData.xlsx')
        se.makeData('AnalysisReport.xlsx')
        print(">>> Data added and saved... <<<")

# ph = PhoneCurry("samsung a50 phonecurry")
# pc = PhoneCurry()
# pc.searchByName("samsung a50 phonecurry", "samsung", "a50")